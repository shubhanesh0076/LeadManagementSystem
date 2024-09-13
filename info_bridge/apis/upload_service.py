# uploads/services.py

import pandas as pd
from django.db import transaction
from info_bridge.models import DataBridge
from leads.models import StudentLeads, ParentsInfo
from locations.models import Address, Country, State, City
from utilities.custom_exceptions import UnexpectedError
from openpyxl import load_workbook


class DataProcessor:

    @staticmethod
    def get_total_rows(file_path):
        workbook = load_workbook(file_path, read_only=True)
        sheet = workbook.active
        return sheet.max_row

    @staticmethod
    def process_excel_in_chunks(file_path):
        chunk_size = 100000
        start_row = 0
        total_rows = DataProcessor.get_total_rows(file_path)
        df_header = pd.read_excel(
            file_path, nrows=0
        )  # Read the header first to get the column names

        columns = df_header.columns
        while True:
            df_chunk = pd.read_excel(
                file_path, skiprows=start_row, nrows=chunk_size, header=0, names=columns
            )
            if df_chunk.empty:
                break
            DataProcessor._process_lead_data(df_chunk)
            start_row += chunk_size
        return total_rows

    @staticmethod
    def process_upload_file(upload_file):
        try:
            file_type = upload_file.name.split(".")[1].upper()

            if file_type == "CSV":
                data = pd.read_csv(upload_file)

            elif file_type == "XLSX":
                return DataProcessor.process_excel_in_chunks(upload_file)

            else:
                raise ValueError("Unsupported file type")

        except Exception as e:
            raise UnexpectedError(message=str(e))

    @staticmethod
    def _process_lead_data(dataframe):

        with transaction.atomic():
            leads_to_create = []
            parents_info = []
            st_location = []

            for index, row in dataframe.iterrows():
                country, _ = Country.objects.get_or_create(
                    name=row.get("country", None).upper()
                )
                state, _ = State.objects.get_or_create(
                    name=row.get("state", None).upper(), country=country
                )
                city, _ = City.objects.get_or_create(
                    name=row.get("city", None).upper(), state=state
                )

                student_lead = StudentLeads(
                    first_name=row.get("first_name"),
                    last_name=row.get("last_name", None),
                    email=row.get("email", None),
                    contact_no=row.get("contact_no", None),
                    alt_contact_no=row.get("alt_contact_no", None),
                    school=(
                        row.get("school", None).upper()
                        if row.get("school", None)
                        else None
                    ),
                )

                student_parents_info = ParentsInfo(
                    lead=student_lead,
                    father_name=row.get("father_name", None),
                    mother_name=row.get("mother_name", None),
                    father_contact_no=row.get("parents_contact_no", None),
                    mother_contact_no=row.get("parents_contact_no", None),
                )

                student_location = Address(
                    lead=student_lead,
                    country=country,
                    state=state,
                    city=city,
                    postal_code=row.get("postal_code", None),
                )

                leads_to_create.append(student_lead)
                parents_info.append(student_parents_info)
                st_location.append(student_location)

            if leads_to_create:
                StudentLeads.objects.bulk_create(leads_to_create)
            if parents_info:
                ParentsInfo.objects.bulk_create(parents_info)
            if st_location:
                Address.objects.bulk_create(st_location)
